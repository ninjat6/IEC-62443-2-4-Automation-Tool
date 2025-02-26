# editors/excel_editor.py
import sys
import os
import json
from PyQt6.QtWidgets import (
    QMainWindow, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QFileDialog, QWidget, QSplitter, QHeaderView, QAbstractItemView,
    QMessageBox, QScrollArea, QPushButton, QSizePolicy, QDialog, QLabel,
    QApplication
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QDropEvent
from openpyxl import load_workbook

# 匯入對話框與自訂元件
from dialogs.edit_dialog import EditDialog
from dialogs.floating_dialog import FloatingDialog
from widgets.draggable_value import DraggableValue
from widgets.removable_block import RemovableBlock
from file_search_module.ui.file_searcher import FileSearcher
from conformity_analysis_module.gui.main_window import ConformityAnalysisWindow

class ExcelEditor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel JSON Drag & Drop Editor")
        self.setGeometry(100, 100, 1200, 800)
        self.showMaximized()

        self.json_data = {}
        self.current_file = None
        self.highlighted_cell = None
        self.info_panel_visible = False  # 預設隱藏資訊面板

        self.conformity_app = None  # 用於管理 ConformityAnalysis 的實例

        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout()

        # 上方按鈕區
        button_layout = QHBoxLayout()
        load_excel_btn = QPushButton("Load Excel")
        load_excel_btn.clicked.connect(self.load_excel)
        button_layout.addWidget(load_excel_btn)

        load_json_btn = QPushButton("Open JSON Elements")
        load_json_btn.clicked.connect(self.open_json_dialog)
        button_layout.addWidget(load_json_btn)

        save_excel_btn = QPushButton("Save Excel")
        save_excel_btn.clicked.connect(self.save_excel)
        button_layout.addWidget(save_excel_btn)

        # 搜尋按鈕
        search_btn = QPushButton("🔍 搜尋文件")
        search_btn.clicked.connect(self.open_search_dialog)
        button_layout.addWidget(search_btn)

        # Conformity Analysis 按鈕
        conformity_btn = QPushButton("Conformity Analysis")
        conformity_btn.clicked.connect(self.open_conformity_analysis)
        button_layout.addWidget(conformity_btn)

        main_layout.addLayout(button_layout)

        # 主區域：左側表格 / 右側資訊面板
        splitter = QSplitter(Qt.Orientation.Horizontal)

        self.table = QTableWidget()
        self.table.setAcceptDrops(True)
        self.table.setDragDropMode(QTableWidget.DragDropMode.DropOnly)
        self.table.viewport().setAcceptDrops(True)
        self.table.dragEnterEvent = self.handle_drag_enter
        self.table.dragMoveEvent = self.handle_drag_move
        self.table.dragLeaveEvent = self.handle_drag_leave
        self.table.dropEvent = self.handle_drop
        self.table.setMouseTracking(True)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setHighlightSections(False)
        self.table.horizontalHeader().sectionClicked.connect(self.display_column_details)
        self.table.cellClicked.connect(self.display_row_details)
        splitter.addWidget(self.table)

        # 資訊面板
        self.info_panel_container = QWidget()
        info_layout = QVBoxLayout()
        info_layout.setContentsMargins(0, 0, 0, 0)

        close_button = QPushButton("✖")
        close_button.setFixedSize(25, 25)
        close_button.setStyleSheet("""
            QPushButton {
                border: none;
                background-color: transparent;
                font-size: 14px;
                color: #555;
            }
            QPushButton:hover {
                color: red;
            }
        """)
        close_button.clicked.connect(self.toggle_info_panel)
        info_layout.addWidget(close_button, alignment=Qt.AlignmentFlag.AlignRight)

        self.info_panel = QScrollArea()
        self.info_panel.setWidgetResizable(True)
        self.info_panel.setStyleSheet("font-size: 14px; padding: 10px; background-color: #f5f5f5; border: 1px solid #ccc;")
        self.info_panel.setMinimumWidth(400)
        info_layout.addWidget(self.info_panel)

        self.info_panel_container.setLayout(info_layout)
        splitter.addWidget(self.info_panel_container)

        # 初始隱藏資訊面板
        self.info_panel_container.hide()
        splitter.setCollapsible(0, True)
        splitter.setCollapsible(1, False)
        splitter.setSizes([800, 400])
        main_layout.addWidget(splitter)

        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

    def open_search_dialog(self):
        """開啟搜尋視窗"""
        self.search_dialog = FileSearcher()
        self.search_dialog.show()

    def open_conformity_analysis(self):
        """開啟 Conformity Analysis 視窗"""
        self.conformity_dialog = ConformityAnalysisWindow()
        self.conformity_dialog.show()

    def toggle_info_panel(self):
        if self.info_panel_visible:
            self.info_panel_container.hide()
        else:
            self.info_panel_container.show()
        self.info_panel_visible = not self.info_panel_visible

    def load_excel(self):
        file_path = os.path.join(os.getcwd(), "resources", "templates", "IEC62443_2_4d_2024-worksheet.xlsx")
        if not os.path.exists(file_path):
            print(f"檔案未找到: {file_path}")
            return

        self.current_file = file_path
        workbook = load_workbook(file_path)
        sheet = workbook.active

        self.hidden_columns = ["Summary Level", "IEC 62443-2-4 Requirement"]
        column_headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        self.visible_columns = [col for col in range(1, sheet.max_column + 1)
                                if column_headers[col - 1] not in self.hidden_columns]
        self.hidden_column_indices = {header: idx for idx, header in enumerate(column_headers)
                                    if header in self.hidden_columns}

        self.table.setRowCount(sheet.max_row - 1)
        self.table.setColumnCount(len(self.visible_columns))

        for col_index, col in enumerate(self.visible_columns):
            self.table.setHorizontalHeaderItem(col_index, QTableWidgetItem(column_headers[col - 1]))

        self.hidden_data = {header: [] for header in self.hidden_columns}
        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=0):
            for col_index, col in enumerate(self.visible_columns):
                cell = row[col - 1]
                item = QTableWidgetItem(str(cell.value) if cell.value else "")
                if col_index == 0:
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled)
                self.table.setItem(row_index, col_index, item)
            for header, idx in self.hidden_column_indices.items():
                cell_value = row[idx].value if idx < len(row) else ""
                self.hidden_data[header].append(cell_value if cell_value else "")

        self.set_column_header_styles()

    def open_json_dialog(self):
        dialog = FloatingDialog(json_data=None, parent=self, multi_level=True)
        dialog.setModal(False)
        dialog.setWindowModality(Qt.WindowModality.NonModal)
        dialog.show()

    def save_excel(self):
        template_path = os.path.join(os.getcwd(), "resources", "templates", "IEC62443_2_4d_2024-worksheet.xlsx")
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            workbook = load_workbook(template_path)
            sheet = workbook.active
            column_map = {
                "Applicant Role(s)": "D",
                "Declared Maturity Level": "E",
                "Conformity Statement": "F",
                "Applicable component": "G",
                "Conformity Evidence": "H",
                "Result - Remarks": "I",
                "Verdict": "J"
            }
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    header_item = self.table.horizontalHeaderItem(col)
                    if not header_item:
                        continue
                    column_name = header_item.text().strip()
                    if column_name not in column_map:
                        continue
                    cell_address = f"{column_map[column_name]}{row + 2}"
                    typed_data = ""
                    item = self.table.item(row, col)
                    if item:
                        typed_data = item.text().strip()
                    block_texts = []
                    cell_widget = self.table.cellWidget(row, col)
                    if cell_widget:
                        layout = cell_widget.layout()
                        if layout:
                            for i in range(layout.count()):
                                block_container = layout.itemAt(i).widget()
                                if block_container:
                                    block_str = block_container.property("block_text")
                                    if block_str:
                                        block_texts.append(block_str)
                    lines_to_write = []
                    if typed_data:
                        lines_to_write.append(typed_data)
                    if block_texts:
                        lines_to_write.extend(block_texts)
                    final_text = "\n\n".join(lines_to_write)
                    sheet[cell_address] = final_text
            workbook.save(file_path)
            QMessageBox.information(self, "Success", "File saved successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save file: {e}")

    def handle_drag_enter(self, event: QDropEvent):
        if event.mimeData().hasText():
            event.acceptProposedAction()
        else:
            event.ignore()

    def handle_drag_move(self, event: QDropEvent):
        pos = event.position().toPoint()
        row = self.table.rowAt(pos.y())
        col = self.table.columnAt(pos.x())
        if col == 0:
            self.clear_highlight()
            event.ignore()
            return
        header_item = self.table.horizontalHeaderItem(col)
        if not header_item:
            event.ignore()
            return
        column_name = header_item.text().strip()
        dragged_text = event.mimeData().text()
        if event.mimeData().data("source_type") == b"DraggableValue":
            allowed_values = {
                "Applicant Role(s)": ["Integrator", "Maintenance Provider"],
                "Declared Maturity Level": ["Level 1 - Initial", "Level 2 - Managed", "Level 3 - Defined", "Level 4 - Improving"],
                "Conformity Statement": ["NULL"],
                "Applicable component": ["Hardware", "Software", "Network Devices", "Data", "Services", "Personnel"],
                "Conformity Evidence": None,
                "Result - Remarks": ["NULL"],
                "Verdict": ["Pass", "Fail", "N/A", "N/E"]
            }
            if column_name not in allowed_values:
                event.ignore()
                return
            if column_name == "Conformity Evidence":
                if not dragged_text.startswith("[") or "TIRI-" not in dragged_text:
                    event.ignore()
                    return
            else:
                allowed_column_values = allowed_values[column_name]
                if allowed_column_values and dragged_text not in allowed_column_values:
                    event.ignore()
                    return
        if row != -1 and col != -1:
            if self.highlighted_cell != (row, col):
                self.clear_highlight()
                self.highlighted_cell = (row, col)
                item = self.table.item(row, col)
                if not item:
                    item = QTableWidgetItem()
                    self.table.setItem(row, col, item)
                item.setBackground(QColor("#D3D3D3"))
            event.acceptProposedAction()
        else:
            event.ignore()

    def handle_drag_leave(self, event: QDropEvent):
        self.clear_highlight()

    def handle_drop(self, event: QDropEvent):
        pos = event.position().toPoint()
        row = self.table.rowAt(pos.y())
        col = self.table.columnAt(pos.x())
        if col == 0:
            event.ignore()
            return
        header_item = self.table.horizontalHeaderItem(col)
        if not header_item:
            event.ignore()
            return
        column_name = header_item.text().strip()
        dragged_text = event.mimeData().text()
        source_type = event.mimeData().data("source_type")
        if source_type == b"DraggableValue":
            allowed_values = {
                "Applicant Role(s)": ["Integrator", "Maintenance Provider"],
                "Declared Maturity Level": ["Level 1 - Initial", "Level 2 - Managed", "Level 3 - Defined", "Level 4 - Improving"],
                "Conformity Statement": ["NULL"],
                "Applicable component": ["Hardware", "Software", "Network Devices", "Data", "Services", "Personnel"],
                "Conformity Evidence": None,
                "Result - Remarks": ["NULL"],
                "Verdict": ["Pass", "Fail", "N/A", "N/E"]
            }
            if column_name not in allowed_values:
                event.ignore()
                return
            if column_name == "Conformity Evidence":
                if not dragged_text.startswith("[") or "TIRI-" not in dragged_text:
                    event.ignore()
                    return
            else:
                allowed_column_values = allowed_values[column_name]
                if allowed_column_values and dragged_text not in allowed_column_values:
                    event.ignore()
                    return
        if row != -1 and col != -1:
            self.add_block_to_table(row, col, dragged_text, source_type)
            self.clear_highlight()
            event.acceptProposedAction()
        else:
            event.ignore()

    def add_block_to_table(self, row, col, text, source_type):
        existing_widget = self.table.cellWidget(row, col)
        existing_data = []
        if existing_widget:
            layout = existing_widget.layout()
            for i in range(layout.count()):
                block_container = layout.itemAt(i).widget()
                if block_container:
                    block_text = block_container.property("block_text")
                    block_type = block_container.property("block_type")
                    if block_text and block_type:
                        existing_data.append((block_text, block_type))
        existing_data.append((text, source_type))
        existing_data = list(set(existing_data))
        existing_data.sort(key=lambda x: x[0])
        if not existing_widget:
            cell_widget = QWidget()
            layout = QVBoxLayout(cell_widget)
            layout.setContentsMargins(5, 5, 5, 5)
            layout.setAlignment(Qt.AlignmentFlag.AlignTop)
            self.table.setCellWidget(row, col, cell_widget)
        else:
            layout = existing_widget.layout()
            while layout.count():
                child = layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()
        for t, tp in existing_data:
            block = self.create_removable_block(t, row, col, tp)
            layout.addWidget(block)
        self.adjust_row_height_based_on_all_cells(row)

    def create_removable_block(self, text, row, col, source_type):
        return RemovableBlock(text, row, col, source_type, editor=self)

    def edit_label_text(self, label, row, col):
        original_text = label.text()
        dialog = EditDialog(original_text, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_text = dialog.new_text
            label.setText(new_text)
            label.setToolTip(new_text)
            self.adjust_row_height_based_on_all_cells(row)

    def remove_block(self, block_container, row, col):
        cell_widget = self.table.cellWidget(row, col)
        if cell_widget:
            layout = cell_widget.layout()
            layout.removeWidget(block_container)
            block_container.deleteLater()
            if layout.count() == 0:
                self.table.removeCellWidget(row, col)
            self.adjust_row_height_based_on_all_cells(row)

    def adjust_row_height_based_on_all_cells(self, row):
        max_height = 0
        for col in range(self.table.columnCount()):
            cell_widget = self.table.cellWidget(row, col)
            if cell_widget:
                layout = cell_widget.layout()
                if layout:
                    cell_height = sum(layout.itemAt(i).widget().sizeHint().height() for i in range(layout.count()))
                    max_height = max(max_height, cell_height)
        self.table.setRowHeight(row, max(max_height + 10, 50))

    def clear_highlight(self):
        if self.highlighted_cell:
            row, col = self.highlighted_cell
            item = self.table.item(row, col)
            if item:
                item.setBackground(QColor("#FFFFFF"))
            self.highlighted_cell = None

    def mouseMoveEvent(self, event):
        pos = event.pos()
        row = self.table.rowAt(pos.y() - self.table.horizontalHeader().height())
        col = self.table.columnAt(pos.x())
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item and not item.isSelected():
                item.setBackground(QColor("#FFFFFF"))
        if row != -1 and col == 0:
            hover_item = self.table.item(row, col)
            if hover_item and not hover_item.isSelected():
                hover_item.setBackground(QColor("#F0F8FF"))
        super().mouseMoveEvent(event)

    def display_row_details(self, row, column):
        selected_item = self.table.item(row, 0)
        if not selected_item:
            self.update_info_panel("Cell Details", [QLabel("No details available.")])
            return
        selected_id = selected_item.text()
        details = []
        for header, data in getattr(self, 'hidden_data', {}).items():
            if row < len(data):
                hidden_label = QLabel(f"<b>{header}:</b> {data[row]}")
                hidden_label.setWordWrap(True)
                hidden_label.setStyleSheet("""
                    QLabel {
                        padding: 10px;
                        background-color: #F9F9F9;
                        color: #333;
                        border: 1px solid #CCC;
                        border-radius: 5px;
                        margin-bottom: 10px;
                    }
                """)
                details.append(hidden_label)
        header_item = self.table.horizontalHeaderItem(column)
        column_name = header_item.text() if header_item else "Cell Block Elements"
        cell_widget = self.table.cellWidget(row, column)
        cell_blocks = []
        if cell_widget:
            layout = cell_widget.layout()
            if layout:
                for i in range(layout.count()):
                    block_container = layout.itemAt(i).widget()
                    if block_container:
                        block_text = block_container.property("block_text")
                        block_type = block_container.property("block_type")
                        if block_text:
                            cell_blocks.append(self.create_editable_block(block_text, block_type, row, column))
        table_item = self.table.item(row, column)
        if table_item and table_item.text().strip():
            input_text = QLabel(f"• {table_item.text().strip()}")
            input_text.setStyleSheet("""
                QLabel {
                    padding: 5px;
                    background-color: #FFF8E1;
                    color: #333333;
                    border: 1px solid #FFD54F;
                    border-radius: 5px;
                    margin-bottom: 5px;
                }
            """)
            cell_blocks.insert(0, input_text)
        if not cell_blocks:
            na_label = QLabel("• N/A")
            na_label.setStyleSheet("""
                QLabel {
                    padding: 5px;
                    background-color: #FFF8E1;
                    color: #333333;
                    border: 1px solid #FFD54F;
                    border-radius: 5px;
                    margin-bottom: 5px;
                }
            """)
            cell_blocks.append(na_label)
        column_title = QLabel(f"<b>{column_name}:</b>")
        column_title.setStyleSheet("""
            QLabel {
                padding: 15px;
                background-color: #003366;
                color: #FFFFFF;
                font-weight: bold;
                font-size: 18px;
                border-radius: 8px 8px 0 0;
            }
        """)
        details.append(column_title)
        details.extend(cell_blocks)
        self.update_info_panel(selected_id, details)

    def create_editable_block(self, block_text, block_type, row, column):
        block = QWidget()
        layout = QHBoxLayout(block)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        delete_button = QPushButton("×", block)
        delete_button.setFixedSize(15, 15)
        delete_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                color: red;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                color: darkred;
            }
        """)
        delete_button.clicked.connect(lambda: self.remove_block_from_panel(block, row, column))
        layout.addWidget(delete_button)
        label = QLabel(block_text, block)
        label.setToolTip(block_text)
        label.setStyleSheet("""
            QLabel {
                padding: 5px;
                background-color: #E8F6FF;
                color: #333333;
                border: 1px solid #80D4FF;
                border-radius: 5px;
                margin-bottom: 5px;
            }
        """)
        layout.addWidget(label)
        if block_type == b"DraggableLabel":
            label.mouseDoubleClickEvent = lambda _: self.edit_block_text(label, row, column)
        return block

    def edit_block_text(self, label, row, column):
        original_text = label.text()
        dialog = EditDialog(original_text, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_text = dialog.new_text
            label.setText(new_text)
            label.setToolTip(new_text)
            cell_widget = self.table.cellWidget(row, column)
            if cell_widget:
                layout = cell_widget.layout()
                for i in range(layout.count()):
                    block_container = layout.itemAt(i).widget()
                    if block_container and block_container.property("block_text") == original_text:
                        block_container.setProperty("block_text", new_text)
                        child_label = block_container.findChild(QLabel)
                        if child_label:
                            child_label.setText(new_text)

    def remove_block_from_panel(self, block, row, column):
        block_text = block.findChild(QLabel).text()
        block.deleteLater()
        cell_widget = self.table.cellWidget(row, column)
        if cell_widget:
            layout = cell_widget.layout()
            if layout:
                for i in range(layout.count() - 1, -1, -1):
                    item = layout.itemAt(i)
                    if item is None:
                        continue
                    block_container = item.widget()
                    if block_container and block_container.property("block_text") == block_text:
                        layout.removeWidget(block_container)
                        block_container.deleteLater()
                if layout.count() == 0:
                    self.table.removeCellWidget(row, column)
            else:
                print(f"No layout found for cell at row {row}, column {column}")
        self.adjust_row_height_based_on_all_cells(row)

    def display_column_details(self, column):
        header_item = self.table.horizontalHeaderItem(column)
        if not header_item:
            self.update_info_panel("Column Details", [QLabel("No details available.")])
            return
        column_name = header_item.text().strip()
        if column_name == "Conformity Evidence":
            self.current_hierarchy = []
            self.load_stage_data([])
        else:
            possible_values = {
                "Applicant Role(s)": ["Integrator", "Maintenance Provider"],
                "Declared Maturity Level": ["Level 1 - Initial", "Level 2 - Managed", "Level 3 - Defined", "Level 4 - Improving"],
                "Applicable component": ["Hardware", "Software", "Network Devices", "Data", "Services", "Personnel"],
                "Verdict": ["Pass", "Fail", "N/A", "N/E"]
            }
            if column_name in possible_values:
                draggable_widgets = [DraggableValue(value, value) for value in possible_values[column_name]]
                self.update_info_panel(f"Column: {column_name}", draggable_widgets)
            else:
                self.update_info_panel("Column Details", [QLabel("No predefined values available for this column.")])

    def load_stage_data(self, hierarchy):
        try:
            json_path = os.path.join(os.getcwd(), "resources", "output_json", "all_documents.json")
            with open(json_path, "r", encoding="utf-8") as f:
                all_data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            self.update_info_panel("Error", [QLabel(f"Failed to load data: {e}")])
            return
        if not hierarchy:
            filtered_data = sorted(set(item.split(" ")[0] for item in all_data if "[" in item and "]" in item))
        else:
            current_stage = hierarchy[-1]
            filtered_data = [item for item in all_data if item.startswith(current_stage)]
        if not filtered_data:
            self.update_info_panel("No Data", [QLabel("No items available for this stage.")])
            return
        if not hierarchy:
            draggable_widgets = []
            for stage in filtered_data:
                draggable = DraggableValue(stage, stage)
                draggable.mousePressEvent = lambda event, s=stage: self.load_stage_data(hierarchy + [s])
                draggable_widgets.append(draggable)
            self.update_info_panel("Conformity Evidence Stages", draggable_widgets)
        else:
            self.render_stage_details(filtered_data, hierarchy)

    def render_stage_details(self, items, hierarchy):
        detail_widgets = []
        if hierarchy:
            back_button = QPushButton("← Back")
            back_button.clicked.connect(lambda: self.load_stage_data(hierarchy[:-1]))
            detail_widgets.append(back_button)
        for item in items:
            full_text = item
            parts = item.split(" ", 2)
            main_part = f"{parts[0]} {parts[1]}" if len(parts) >= 2 else item
            tooltip_text = parts[2] if len(parts) > 2 else ""
            draggable = DraggableValue(main_part, full_text)
            draggable.setToolTip(tooltip_text)
            detail_widgets.append(draggable)
        self.update_info_panel("Stage Details", detail_widgets)

    def update_info_panel(self, title, content_widgets):
        if not self.info_panel_visible:
            self.info_panel_container.show()
            self.info_panel_visible = True
        self.info_panel.setStyleSheet("""
            QWidget {
                background-color: #FFFFFF;
                border: 1px solid #CCE7FF;
                border-radius: 8px;
            }
        """)
        panel_layout = QVBoxLayout()
        panel_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        title_label = QLabel(title)
        title_label.setStyleSheet("""
            QLabel {
                padding: 15px;
                background-color: #003366;
                color: #FFFFFF;
                font-weight: bold;
                font-size: 18px;
                border-radius: 8px 8px 0 0;
            }
        """)
        panel_layout.addWidget(title_label)
        for widget in content_widgets:
            if isinstance(widget, DraggableValue):
                widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            elif isinstance(widget, QLabel):
                widget.setWordWrap(True)
                widget.setStyleSheet("""
                    QLabel {
                        padding: 10px;
                        background-color: #F8FBFF;
                        color: #333333;
                        border: 1px solid #CCE7FF;
                        border-radius: 5px;
                        margin-bottom: 10px;
                        font-size: 14px;
                    }
                """)
            widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            panel_layout.addWidget(widget)
        container = QWidget()
        container.setLayout(panel_layout)
        self.info_panel.setWidget(container)
        self.info_panel.setContentsMargins(20, 20, 20, 20)

    def set_column_header_styles(self):
        for col in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(col)
            if header_item:
                header_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                font = header_item.font()
                font.setBold(True)
                header_item.setFont(font)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, 'info_panel') and self.info_panel.widget():
            panel_width = self.info_panel.width() - 30
            block_height = 50
            layout = self.info_panel.widget().layout()
            if layout:
                for i in range(layout.count()):
                    widget = layout.itemAt(i).widget()
                    if widget and hasattr(widget, "set_size"):
                        widget.set_size(panel_width, block_height)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    editor = ExcelEditor()
    editor.show()
    sys.exit(app.exec())