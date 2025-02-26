# utils/file_processor.py

from docx import Document
from openpyxl import load_workbook
import json
import os
import re
import subprocess
import tempfile

class FileProcessor:
    """處理檔案並分割中英文內容的工具類"""

    @staticmethod
    def clean_line_chinese(line: str) -> str:
        """
        清理單行內容，保留：
          1. 中文（含數字）
          2. 全部字母皆大寫、且長度 >= 2 的英文單字（例如 TIRI、CPU、ABC）
          3. TIRI/IEC 編號（如 TIRI-E-1-0-001, IEC62443-2-4:2015）
          4. 符號 - : / 、（與中文、數字並存時可保留）
        """
        line = line.strip()
        if not line:
            return ""

        protected = {}

        # 1) 保護 TIRI 文件編號
        doc_nums = re.findall(r'TIRI-[A-Z]-\d-\d-\d{3}', line)
        for i, num in enumerate(doc_nums):
            key = f"{{DOC{i}}}"
            protected[key] = num
            line = line.replace(num, key)

        # 2) 保護 IEC 編號
        iec_pat = r'(IEC\s*\d+(?:-\d+)*(?::\d+)?(?:-\d+)?)'
        iec_nums = re.findall(iec_pat, line)
        for i, num in enumerate(iec_nums):
            key = f"{{IEC{i}}}"
            protected[key] = num.replace(' ', '')
            line = line.replace(num, key)

        # 3) 若冒號前無中文，則只保留冒號後（原需求）
        if re.search(r'[：:]', line):
            parts = re.split(r'[：:]', line, 1)
            if len(parts) == 2:
                if not re.search(r'[\u4e00-\u9fff]', parts[0]):
                    line = parts[1].strip()
                else:
                    line = parts[0].strip() + "：" + parts[1].strip()

        # 4) 用正則一次性擷取 token
        pattern = r'[\u4e00-\u9fff0-9\-\:\/、]+|[A-Z]{2,}|\{DOC\d+\}|\{IEC\d+\}'
        tokens = re.findall(pattern, line)
        if not tokens:
            return ""

        new_line = " ".join(tokens)

        # 5) 還原 placeholder
        for ph, val in protected.items():
            new_line = new_line.replace(ph, val)

        return new_line.strip()

    @staticmethod
    def clean_line_english(line: str) -> str:
        """
        清理單行內容，保留英文（含大小寫）、數字、及 TIRI/IEC 編號；移除中文。
        """
        line = line.strip()
        if not line:
            return ""

        # 若沒任何英文字母/數字/TIRI/IEC，就視為非英文
        if not re.search(r'[a-zA-Z0-9]|TIRI-|IEC', line):
            return ""

        protected = {}

        # 1) 保護 TIRI 文件編號
        doc_nums = re.findall(r'TIRI-[A-Z]-\d-\d-\d{3}', line)
        for i, num in enumerate(doc_nums):
            key = f"{{DOCENG{i}}}"
            protected[key] = num
            line = line.replace(num, key)

        # 2) 保護 IEC
        iec_nums = re.findall(r'IEC\s*\d+(?:-\d+)*(?:-\d+)?', line)
        for i, num in enumerate(iec_nums):
            key = f"{{IECENG{i}}}"
            protected[key] = num.replace(' ', '')
            line = line.replace(num, key)

        # 移除所有中文（含中文標點）
        line = re.sub(r'[\u4e00-\u9fff]', '', line)
        # 轉換常見中文標點
        line = line.replace('：', ':').replace('，', ',')
        line = line.replace('（', '(').replace('）', ')')

        # 合併多重空白
        line = re.sub(r'\s+', ' ', line).strip()

        # 還原 placeholder
        for placeholder, original_str in protected.items():
            line = line.replace(placeholder, original_str)

        if not re.search(r'[a-zA-Z0-9]|TIRI-|IEC', line):
            return ""

        return line

    @staticmethod
    def split_text_advanced(text: str):
        """
        將整份文本逐行分割，分別產生「中文清單」與「英文清單」。
        """
        lines = text.split('\n')
        chinese_content = []
        english_content = []

        for raw_line in lines:
            line = raw_line.strip()
            if not line:
                continue

            c_line = FileProcessor.clean_line_chinese(line)
            if c_line:
                chinese_content.append(c_line)

            e_line = FileProcessor.clean_line_english(line)
            if e_line:
                english_content.append(e_line)

        # 去重且保序
        def deduplicate_preserve_order(items):
            seen = set()
            result = []
            for x in items:
                if x not in seen:
                    seen.add(x)
                    result.append(x)
            return result

        chinese_content = deduplicate_preserve_order(chinese_content)
        english_content = deduplicate_preserve_order(english_content)

        return chinese_content, english_content

    @staticmethod
    def split_docx_to_json(input_path, output_chinese_path, output_english_path):
        """處理 Word 文件 (.docx) -> 生成中英文 JSON"""
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"找不到輸入檔案：{input_path}")

        try:
            doc = Document(input_path)
            # 收集所有段落文字並以換行符連成大字串
            text = "\n".join(para.text for para in doc.paragraphs if para.text.strip())

            chinese_content, english_content = FileProcessor.split_text_advanced(text)

            os.makedirs(os.path.dirname(output_chinese_path), exist_ok=True)
            os.makedirs(os.path.dirname(output_english_path), exist_ok=True)

            with open(output_chinese_path, 'w', encoding='utf-8') as ch_file:
                json.dump(chinese_content, ch_file, ensure_ascii=False, indent=4)

            with open(output_english_path, 'w', encoding='utf-8') as en_file:
                json.dump(english_content, en_file, ensure_ascii=False, indent=4)

            return {
                "chinese_json": output_chinese_path,
                "english_json": output_english_path
            }

        except Exception as e:
            raise Exception(f"處理 DOCX 檔案時發生錯誤：{str(e)}")

    @staticmethod
    def split_xlsx_to_json(input_path, output_chinese_path, output_english_path):
        """處理 Excel 文件 -> 生成中英文 JSON"""
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"找不到輸入檔案：{input_path}")

        try:
            workbook = load_workbook(input_path)
            sheet = workbook.active

            text_lines = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    line_str = " ".join(str(cell) for cell in row if cell is not None)
                    text_lines.append(line_str)

            text = "\n".join(text_lines)
            chinese_content, english_content = FileProcessor.split_text_advanced(text)

            os.makedirs(os.path.dirname(output_chinese_path), exist_ok=True)
            os.makedirs(os.path.dirname(output_english_path), exist_ok=True)

            with open(output_chinese_path, 'w', encoding='utf-8') as ch_file:
                json.dump(chinese_content, ch_file, ensure_ascii=False, indent=4)

            with open(output_english_path, 'w', encoding='utf-8') as en_file:
                json.dump(english_content, en_file, ensure_ascii=False, indent=4)

            return {
                "chinese_json": output_chinese_path,
                "english_json": output_english_path
            }

        except Exception as e:
            raise Exception(f"處理 XLSX 檔案時發生錯誤：{str(e)}")

    @staticmethod
    def convert_doc_to_docx(input_path):
        """
        若為 .doc 檔，透過 LibreOffice（或 unoconv）將其轉換為 .docx。
        回傳轉檔後的 .docx 路徑。若失敗則拋出例外。
        """
        base, _ = os.path.splitext(input_path)
        temp_docx = base + ".temp.docx"

        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                subprocess.run([
                    "soffice", "--headless", "--convert-to", "docx", 
                    "--outdir", tmpdir, input_path
                ], check=True)
                # LibreOffice 轉檔後檔名與原檔相同但副檔名改為 docx
                doc_name = os.path.basename(base) + ".docx"
                converted_path = os.path.join(tmpdir, doc_name)
                if not os.path.exists(converted_path):
                    raise FileNotFoundError(f"LibreOffice 轉檔失敗，找不到: {converted_path}")

                # 複製轉檔結果到 temp_docx
                with open(converted_path, "rb") as src, open(temp_docx, "wb") as dst:
                    dst.write(src.read())

            if not os.path.exists(temp_docx):
                raise FileNotFoundError(f"找不到轉檔後的檔案: {temp_docx}")

            return temp_docx

        except subprocess.CalledProcessError as e:
            raise Exception(f".doc 轉 .docx 失敗: {str(e)}")

    @staticmethod
    def process_file(input_path, output_dir):
        """
        處理檔案的主要方法：
        1. 檢查檔案存在
        2. 判斷副檔名
        3. 呼叫對應方法（doc/docx / xlsx），輸出中文與英文 JSON
        """
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"找不到輸入檔案：{input_path}")

        file_name = os.path.splitext(os.path.basename(input_path))[0]
        output_chinese_path = os.path.join(output_dir, f"{file_name}_chinese.json")
        output_english_path = os.path.join(output_dir, f"{file_name}_english.json")

        ext = os.path.splitext(input_path)[1].lower()
        if ext in ['.doc', '.docx']:
            # 若為 .doc，先轉檔再處理
            if ext == '.doc':
                converted_path = FileProcessor.convert_doc_to_docx(input_path)
                return FileProcessor.split_docx_to_json(converted_path,
                                                        output_chinese_path,
                                                        output_english_path)
            else:
                # .docx 直接處理
                return FileProcessor.split_docx_to_json(input_path,
                                                        output_chinese_path,
                                                        output_english_path)
        elif ext == '.xlsx':
            return FileProcessor.split_xlsx_to_json(input_path,
                                                    output_chinese_path,
                                                    output_english_path)
        else:
            raise ValueError("不支援的檔案格式。僅支援 .doc, .docx 和 .xlsx 格式。")


if __name__ == "__main__":
    """
    若要從命令列執行:
    python file_processor.py <input_file_path> <output_dir>
    """
    import sys
    if len(sys.argv) < 3:
        print("Usage: python file_processor.py <input_file_path> <output_dir>")
        sys.exit(1)

    in_path = sys.argv[1]
    out_dir = sys.argv[2]

    try:
        result = FileProcessor.process_file(in_path, out_dir)
        print("Process success:", result)
    except Exception as e:
        print("Error:", e)
        sys.exit(1)
