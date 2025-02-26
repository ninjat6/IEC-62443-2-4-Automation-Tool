import os
import re
import chardet
import docx
import openpyxl
from PyPDF2 import PdfReader
from PyQt6.QtCore import QObject, pyqtSignal

class SearchWorker(QObject):
    progress_update = pyqtSignal(int)
    search_finished = pyqtSignal(str)
    error_occurred = pyqtSignal(str, str)
    file_matches_found = pyqtSignal(str, list)  # (檔案路徑, [符合的行])
    
    def __init__(self, folder, keyword):
        super().__init__()
        self.folder = folder
        self.keyword = keyword
        self.cancel_search_flag = False
        self.total_files = 0
        self.current_progress = 0
    
    def start_search(self):
        self.total_files = self.count_target_files(self.folder)
        if self.total_files == 0:
            self.search_finished.emit("找不到可搜尋的檔案！")
            return
        
        for root, _, files in os.walk(self.folder):
            if self.cancel_search_flag:
                self.search_finished.emit("搜尋已取消！")
                return
            
            for file in files:
                if self.cancel_search_flag:
                    self.search_finished.emit("搜尋已取消！")
                    return
                
                file_path = os.path.join(root, file)
                ext = file.lower().rsplit('.', 1)[-1]
                
                try:
                    if ext == "txt":
                        self.process_text_file(file_path, self.keyword)
                    elif ext == "pdf":
                        self.process_pdf_file(file_path, self.keyword)
                    elif ext == "docx":
                        self.process_docx_file(file_path, self.keyword)
                    elif ext == "xlsx":
                        self.process_xlsx_file(file_path, self.keyword)
                    elif ext in ("html", "htm"):
                        self.process_text_file(file_path, self.keyword)
                except Exception as e:
                    self.error_occurred.emit("錯誤", f"處理檔案 {file_path} 時發生問題: {e}")
                
                self.current_progress += 1
                self.progress_update.emit(self.current_progress)
        
        self.search_finished.emit("搜尋完成！")
    
    def cancel_search(self):
        self.cancel_search_flag = True
    
    def count_target_files(self, folder):
        count = 0
        for root, _, files in os.walk(folder):
            for file in files:
                if file.lower().endswith((".txt", ".pdf", ".docx", ".xlsx", ".html", ".htm")):
                    count += 1
        return count
    
    def process_text_file(self, file_path, keyword):
        try:
            with open(file_path, "rb") as f:
                raw_data = f.read()
            result = chardet.detect(raw_data)
            detected_encoding = result["encoding"] or "utf-8"
            text = raw_data.decode(detected_encoding, errors="replace")
        except Exception as e:
            self.error_occurred.emit("錯誤", f"無法讀取文字檔 {file_path}: {e}")
            return
        
        lines = text.splitlines()
        matches = []
        for line in lines:
            if re.search(re.escape(keyword), line, re.IGNORECASE):
                matches.append(line.strip())
        if matches:
            self.file_matches_found.emit(file_path, matches)
    
    def process_pdf_file(self, file_path, keyword):
        lines = []
        try:
            reader = PdfReader(file_path)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    lines.extend(text.split('\n'))
        except Exception as e:
            self.error_occurred.emit("錯誤", f"無法處理 PDF {file_path}: {e}")
            return
        
        matches = []
        for line in lines:
            if re.search(re.escape(keyword), line, re.IGNORECASE):
                matches.append(line.strip())
        if matches:
            self.file_matches_found.emit(file_path, matches)
    
    def process_docx_file(self, file_path, keyword):
        try:
            doc = docx.Document(file_path)
            lines = [para.text for para in doc.paragraphs]
        except Exception as e:
            self.error_occurred.emit("錯誤", f"無法處理 Word 文件 {file_path}: {e}")
            return
        
        matches = []
        for line in lines:
            if re.search(re.escape(keyword), line, re.IGNORECASE):
                matches.append(line.strip())
        if matches:
            self.file_matches_found.emit(file_path, matches)
    
    def process_xlsx_file(self, file_path, keyword):
        lines = []
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join(str(cell) for cell in row if cell is not None)
                    lines.append(row_text)
        except Exception as e:
            self.error_occurred.emit("錯誤", f"無法處理 Excel 文件 {file_path}: {e}")
            return
        
        matches = []
        for line in lines:
            if re.search(re.escape(keyword), line, re.IGNORECASE):
                matches.append(line.strip())
        if matches:
            self.file_matches_found.emit(file_path, matches)
