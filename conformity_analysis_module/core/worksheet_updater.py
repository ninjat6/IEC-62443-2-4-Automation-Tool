import os
import json
import re
import shutil
import openpyxl
import tkinter as tk
from tkinter import filedialog
from conformity_analysis_module.utils.docx_section_extractor import DocxSectionExtractor
from conformity_analysis_module.utils.logger import logger
from conformity_analysis_module.config import Config

class WorksheetUpdater:
    @staticmethod
    def update_worksheet():
        analysis_file = Config.ANALYSIS_OUTPUT
        template_file = Config.WORKSHEET_FILE
        print(analysis_file)
        if not os.path.exists(analysis_file):
            return False, "analysis_results.json 不存在，請先執行分析"
        if not os.path.exists(template_file):
            print(template_file)
            return False, "找不到模板檔案 IEC62443_2_4d_2024-worksheet.xlsx"

        try:
            with open(analysis_file, 'r', encoding='utf-8') as f:
                analysis_results = json.load(f)
        except Exception as e:
            logger.error(f"讀取 {analysis_file} 時發生錯誤: {e}")
            return False, "讀取 analysis_results.json 時發生錯誤"

        # 複製模板檔案到臨時位置
        temp_dir = os.path.join(Config.ROOT_DIR, "temp")
        os.makedirs(temp_dir, exist_ok=True)
        temp_file = os.path.join(temp_dir, "IEC62443_2_4d_filled.xlsx")
        shutil.copy(template_file, temp_file)

        try:
            wb = openpyxl.load_workbook(temp_file)
            ws = wb.active

            header = {str(ws.cell(row=1, column=col).value).strip().lower(): col 
                      for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value}
            required_headers = ["iec 62443-2-4 id", "conformity statement", "conformity evidence"]
            for h in required_headers:
                if h not in header:
                    return False, f"Worksheet 缺少必要的欄位: {h}"

            id_col, stmt_col, evi_col = (header[h] for h in required_headers)

            def normalize_req(r):
                return re.sub(r"\s+", "", str(r).upper())

            results_by_requirement = {}
            for entry in analysis_results:
                req = normalize_req(entry.get("requirement", ""))
                if req:
                    results_by_requirement.setdefault(req, []).append(entry)

            grouped_by_source = {}
            for req, entries in results_by_requirement.items():
                source_order = []
                source_best = {}
                for entry in entries:
                    src = entry["source_file"]
                    sim = entry.get("similarity", 0)
                    if src not in source_best:
                        source_best[src] = entry
                        source_order.append(src)
                    elif sim > source_best[src]["similarity"]:
                        source_best[src] = entry
                grouped_by_source[req] = [source_best[s] for s in source_order]

            for row in range(2, ws.max_row + 1):
                req_id = normalize_req(ws.cell(row=row, column=id_col).value)
                if req_id and req_id in results_by_requirement:
                    all_entries = results_by_requirement[req_id]
                    snippets_with_source = []
                    for entry in all_entries:
                        src = entry["source_file"]
                        snippet = entry["snippet"]
                        filename_no_ext = re.sub(r'[\u4e00-\u9fa5]+', '', os.path.splitext(os.path.basename(src))[0])
                        section_str = ""
                        if src.lower().endswith('.docx'):
                            try:
                                extractor = DocxSectionExtractor(src)
                                section_str = extractor.get_section_number(snippet)
                            except Exception as e:
                                logger.error(f"取得 section 編號失敗，檔案 {src}: {e}")
                                section_str = "(無編號)"
                        snippet_text = f"{filename_no_ext}\n Section {section_str}described that {snippet}"
                        snippets_with_source.append(snippet_text)

                    ws.cell(row=row, column=stmt_col, value="\n\n".join(snippets_with_source))

                    evidence_entries = grouped_by_source.get(req_id, [])
                    filenames = [os.path.splitext(os.path.basename(e["source_file"]))[0] for e in evidence_entries]
                    ws.cell(row=row, column=evi_col, value="\n\n".join(filenames))

            # 讓使用者選擇存檔位置
            root = tk.Tk()
            root.withdraw()  # 隱藏 Tkinter 視窗
            save_path = filedialog.asksaveasfilename(
                title="選擇儲存 Excel 檔案",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="IEC62443_2_4d_filled.xlsx"
            )

            if save_path:
                wb.save(save_path)
                logger.info(f"Worksheet 更新完成，已儲存至 {save_path}")
                return True, f"Worksheet 更新完成，檔案已儲存至 {save_path}"
            else:
                logger.warning("使用者取消存檔")
                return False, "使用者取消存檔"

        except Exception as e:
            logger.error(f"更新 Worksheet 時發生錯誤: {e}")
            return False, "更新 Worksheet 時發生錯誤"
